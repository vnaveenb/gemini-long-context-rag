"""XLSX text extraction using openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.logger import get_logger
from src.models.document import (
    DocumentFormat,
    DocumentMetadata,
    ExtractedContent,
    PageContent,
    Section,
)

logger = get_logger(__name__)


def extract_xlsx(file_path: str | Path) -> ExtractedContent:
    """Extract text, sections, and metadata from an XLSX file."""
    file_path = Path(file_path)
    logger.info("Extracting XLSX", path=str(file_path))

    errors: list[str] = []

    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Failed to open XLSX", path=str(file_path), error=str(exc))
        return ExtractedContent(
            filename=file_path.name,
            format=DocumentFormat.XLSX,
            extraction_errors=[f"Failed to open XLSX: {exc}"],
        )

    pages: list[PageContent] = []
    sections: list[Section] = []
    all_text_parts: list[str] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        try:
            ws = wb[sheet_name]
            rows_text: list[str] = []
            header_row: list[str] = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cell_values = [str(c) if c is not None else "" for c in row]
                # Skip fully empty rows
                if not any(v.strip() for v in cell_values):
                    continue

                row_text = " | ".join(cell_values)
                rows_text.append(row_text)

                if row_idx == 1:
                    header_row = cell_values

            sheet_text = f"[Sheet: {sheet_name}]\n"
            if header_row:
                sheet_text += f"Headers: {' | '.join(header_row)}\n"
            sheet_text += "\n".join(rows_text)

            pages.append(
                PageContent(
                    page_number=sheet_idx,
                    text=sheet_text,
                    metadata={"sheet_name": sheet_name, "row_count": len(rows_text)},
                )
            )
            all_text_parts.append(sheet_text)

            sections.append(
                Section(
                    title=sheet_name,
                    level=1,
                    page_start=sheet_idx,
                    page_end=sheet_idx,
                    content=sheet_text,
                )
            )
        except Exception as exc:
            errors.append(f"Sheet '{sheet_name}': {exc}")
            logger.warning("Error extracting sheet", sheet=sheet_name, error=str(exc))

    wb.close()

    raw_text = "\n\n".join(all_text_parts)

    metadata = DocumentMetadata(
        page_count=len(pages),
        word_count=len(raw_text.split()),
    )

    content = ExtractedContent(
        filename=file_path.name,
        format=DocumentFormat.XLSX,
        metadata=metadata,
        pages=pages,
        sections=sections,
        raw_text=raw_text,
        extraction_errors=errors,
    )
    logger.info(
        "XLSX extraction complete",
        sheets=len(pages),
        words=metadata.word_count,
    )
    return content
